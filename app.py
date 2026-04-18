import os
import uuid
import time
from flask import Flask, render_template, request, send_from_directory
from openpyxl import Workbook, load_workbook

from ML.inference import load_my_model, predict_video

app = Flask(__name__)

# ----------------------------
# 1) Upload folder
# ----------------------------
UPLOAD_FOLDER = "uploads"
DEBUG_FRAMES_FOLDER = "debug_frames"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DEBUG_FRAMES_FOLDER, exist_ok=True)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# ----------------------------
# 2) Load model once
# ----------------------------
MODEL_PATH = r"C:\Users\kudus\Desktop\try out\saved_models\best_model_tryout.pth"
my_model = load_my_model(MODEL_PATH)

# ----------------------------
# 3) Excel results file
# ----------------------------
EXCEL_PATH = r"C:\Users\kudus\Desktop\try out\test_results\tryout_model_results.xlsx"
os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)


# ----------------------------
# 4) Confidence helpers
# ----------------------------
def get_confidence_level(confidence):
    if confidence < 65:
        return "Lower Confidence"
    elif confidence < 80:
        return "Good Confidence"
    else:
        return "High Confidence"


def generate_prediction_feedback(label, confidence, prob_real, prob_fake):
    label = label.upper()
    confidence_level = get_confidence_level(confidence)

    if label == "FAKE":
        if confidence < 65:
            return (
                f"The video was classified as FAKE with {confidence_level.lower()}. "
                f"This indicates that the model detected more visual patterns associated with manipulated content "
                f"than authentic footage, although the distinction was not especially strong. "
                f"For this reason, the result should be interpreted with caution."
            )
        elif confidence < 80:
            return (
                f"The video was classified as FAKE with {confidence_level.lower()}. "
                f"The analysed frames showed several characteristics that were more consistent with altered or synthetic content "
                f"than with genuine video, providing a solid basis for the prediction."
            )
        else:
            return (
                f"The video was classified as FAKE with {confidence_level.lower()}. "
                f"This suggests that the selected frames contained strong visual characteristics more commonly associated "
                f"with manipulated or synthetic media than authentic content."
            )

    elif label == "REAL":
        if confidence < 65:
            return (
                f"The video was classified as REAL with {confidence_level.lower()}. "
                f"This means the model leaned toward authentic content, but the distinction between real and fake was not particularly strong. "
                f"The result should therefore be treated as a cautious indication rather than a highly certain conclusion."
            )
        elif confidence < 80:
            return (
                f"The video was classified as REAL with {confidence_level.lower()}. "
                f"The analysed frames appeared more consistent with genuine video characteristics than manipulated content, "
                f"providing a reliable basis for the prediction."
            )
        else:
            return (
                f"The video was classified as REAL with {confidence_level.lower()}. "
                f"This indicates that the selected frames showed strong visual patterns consistent with authentic, non-manipulated video content."
            )

    return "The model could not generate a detailed explanation for this prediction."


def generate_result_summary(label, confidence):
    label = label.upper()

    if label == "FAKE":
        if confidence < 65:
            return "The model leaned slightly toward FAKE."
        elif confidence < 80:
            return "The model leaned clearly toward FAKE."
        else:
            return "The model leaned strongly toward FAKE."

    if label == "REAL":
        if confidence < 65:
            return "The model leaned slightly toward REAL."
        elif confidence < 80:
            return "The model leaned clearly toward REAL."
        else:
            return "The model leaned strongly toward REAL."

    return "No summary available."


def format_file_size(num_bytes):
    if num_bytes < 1024:
        return f"{num_bytes} B"
    elif num_bytes < 1024 * 1024:
        return f"{num_bytes / 1024:.2f} KB"
    elif num_bytes < 1024 * 1024 * 1024:
        return f"{num_bytes / (1024 * 1024):.2f} MB"
    else:
        return f"{num_bytes / (1024 * 1024 * 1024):.2f} GB"


# ----------------------------
# 5) Home
# ----------------------------
@app.route("/", methods=["GET"])
def home():
    return render_template("index.html")


# ----------------------------
# Helper: safe extension check
# ----------------------------
def allowed_video(filename: str) -> bool:
    filename = filename.lower()
    return (
        filename.endswith(".mp4")
        or filename.endswith(".mov")
        or filename.endswith(".avi")
        or filename.endswith(".mkv")
    )


# ----------------------------
# Serve uploaded videos
# ----------------------------
@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


# ----------------------------
# Serve saved debug frames
# ----------------------------
@app.route("/debug_frames/<path:filename>")
def debug_file(filename):
    return send_from_directory(DEBUG_FRAMES_FOLDER, filename)


# ----------------------------
# 6) Predict
# ----------------------------
@app.route("/predict", methods=["POST"])
def predict():
    if "file" not in request.files:
        return render_template("index.html", error="No file uploaded.")

    file = request.files["file"]

    if file.filename == "":
        return render_template("index.html", error="No file selected.")

    if not allowed_video(file.filename):
        return render_template("index.html", error="Please upload a video file.")

    original_filename = file.filename
    file_ext = os.path.splitext(original_filename)[1].lower()
    file_bytes = request.content_length if request.content_length else 0

    # Save with a unique name
    ext = os.path.splitext(file.filename)[1].lower()
    unique_name = f"{uuid.uuid4().hex}{ext}"
    save_path = os.path.join(app.config["UPLOAD_FOLDER"], unique_name)
    file.save(save_path)

    actual_file_size = os.path.getsize(save_path)

    # Make a unique debug folder per upload
    per_video_debug = os.path.join(DEBUG_FRAMES_FOLDER, os.path.splitext(unique_name)[0])
    os.makedirs(per_video_debug, exist_ok=True)

    # Prediction
    start_time = time.time()

    label, confidence, probs = predict_video(
        my_model,
        save_path,
        frame_limit=10,
        debug_dir=per_video_debug
    )

    processing_time = time.time() - start_time

    if label == "ERROR":
        return render_template("index.html", error="Could not read this video.")

    # Confidence and explanation
    confidence_level = get_confidence_level(confidence)
    prediction_feedback = generate_prediction_feedback(
        label,
        confidence,
        probs[0] * 100,
        probs[1] * 100
    )
    result_summary = generate_result_summary(label, confidence)

    # Collect saved frame file names
    frame_files = []
    if os.path.exists(per_video_debug):
        frame_files = sorted(
            [f for f in os.listdir(per_video_debug) if f.lower().endswith((".jpg", ".jpeg", ".png"))]
        )

    frame_count = len(frame_files)

    # ----------------------------
    # Save result to Excel
    # ----------------------------
    filename_upper = original_filename.upper()

    if "FAKE" in filename_upper:
        actual_label = "FAKE"
    elif "REAL" in filename_upper:
        actual_label = "REAL"
    else:
        actual_label = "UNKNOWN"

    if actual_label == "UNKNOWN":
        result_status = "Unknown"
    else:
        result_status = "Correct" if label.upper() == actual_label else "Wrong"

    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append([
            "video_name",
            "actual_label",
            "predicted_label",
            "confidence",
            "real_probability",
            "fake_probability",
            "result_status"
        ])
        wb.save(EXCEL_PATH)

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    ws.append([
        original_filename,
        actual_label,
        label.upper(),
        round(confidence, 2),
        round(probs[0] * 100, 2),
        round(probs[1] * 100, 2),
        result_status
    ])

    wb.save(EXCEL_PATH)

    return render_template(
        "index.html",
        prediction=label.upper(),
        confidence=f"{confidence:.2f}",
        confidence_level=confidence_level,
        prob_real=f"{probs[0] * 100:.2f}",
        prob_fake=f"{probs[1] * 100:.2f}",
        uploaded_video=unique_name,
        debug_folder=os.path.splitext(unique_name)[0],
        frame_files=frame_files,
        frame_count=frame_count,
        prediction_feedback=prediction_feedback,
        result_summary=result_summary,
        processing_time=f"{processing_time:.2f}",
        uploaded_filename=original_filename,
        uploaded_filetype=file_ext.upper().replace(".", ""),
        uploaded_filesize=format_file_size(actual_file_size),
        status_saved="Saved to Excel"
    )


if __name__ == "__main__":
    app.run(debug=True)